"""
Batch processing service for handling bulk hallucination detection.
"""

import asyncio
import csv
import json
import uuid
import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional, AsyncGenerator
from pathlib import Path
import logging

from src.models.batch_schemas import (
    BatchJob, BatchStatus, BatchFileFormat, BatchInputRow, BatchResultRow,
    BatchProgressUpdate, BatchUploadRequest
)
from src.models.schemas import AgentTestRequest
from src.judges.ensemble_judge import EnsembleJudge

logger = logging.getLogger(__name__)


class BatchProcessor:
    """Handles batch processing of hallucination detection jobs."""
    
    def __init__(self, claude_api_key: str, storage_dir: str = "batch_storage"):
        self.judge = EnsembleJudge(claude_api_key)
        self.storage_dir = Path(storage_dir)
        self.storage_dir.mkdir(exist_ok=True)
        
        # In-memory job tracking (in production, use Redis/database)
        self.jobs: Dict[str, BatchJob] = {}
        self.active_tasks: Dict[str, asyncio.Task] = {}
        
        # Progress callbacks for real-time updates
        self.progress_callbacks: List[callable] = []
    
    def add_progress_callback(self, callback: callable):
        """Add callback for progress updates."""
        self.progress_callbacks.append(callback)
    
    async def _notify_progress(self, update: BatchProgressUpdate):
        """Notify all progress callbacks."""
        for callback in self.progress_callbacks:
            try:
                await callback(update)
            except Exception as e:
                logger.error(f"Error in progress callback: {e}")
    
    async def upload_file(self, file_content: bytes, upload_request: BatchUploadRequest) -> BatchJob:
        """Upload and validate a batch file."""
        job_id = str(uuid.uuid4())
        timestamp = datetime.utcnow()
        
        # Save uploaded file
        input_file_path = self.storage_dir / f"{job_id}_input.{upload_request.file_format.value}"
        with open(input_file_path, 'wb') as f:
            f.write(file_content)
        
        # Parse and validate file
        try:
            rows = await self._parse_file(input_file_path, upload_request)
            total_rows = len(rows)
            
            if total_rows == 0:
                raise ValueError("File contains no valid data rows")
            
            # Create job
            job = BatchJob(
                job_id=job_id,
                filename=upload_request.filename,
                file_format=upload_request.file_format,
                status=BatchStatus.PENDING,
                total_rows=total_rows,
                created_at=timestamp,
                input_file_path=str(input_file_path)
            )
            
            self.jobs[job_id] = job
            logger.info(f"Created batch job {job_id} with {total_rows} rows")
            
            return job
            
        except Exception as e:
            # Clean up file on error
            if input_file_path.exists():
                input_file_path.unlink()
            raise ValueError(f"Failed to parse file: {str(e)}")
    
    async def _parse_file(self, file_path: Path, upload_request: BatchUploadRequest) -> List[BatchInputRow]:
        """Parse uploaded file into BatchInputRow objects."""
        rows = []
        
        if upload_request.file_format == BatchFileFormat.CSV:
            # Parse CSV
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f) if upload_request.has_headers else csv.reader(f)
                
                for i, row in enumerate(reader):
                    try:
                        if upload_request.has_headers:
                            # Use column mappings
                            agent_output = row.get(upload_request.agent_output_column, "").strip()
                            if not agent_output:
                                continue
                                
                            batch_row = BatchInputRow(
                                agent_output=agent_output,
                                ground_truth=row.get(upload_request.ground_truth_column) if upload_request.ground_truth_column else None,
                                query=row.get(upload_request.query_column) if upload_request.query_column else None,
                                agent_id=row.get(upload_request.agent_id_column) if upload_request.agent_id_column else None,
                                metadata={"row_number": i + 1, "original_row": dict(row)}
                            )
                        else:
                            # Assume first column is agent_output
                            if len(row) == 0 or not row[0].strip():
                                continue
                            batch_row = BatchInputRow(
                                agent_output=row[0].strip(),
                                ground_truth=row[1] if len(row) > 1 else None,
                                query=row[2] if len(row) > 2 else None,
                                agent_id=row[3] if len(row) > 3 else None,
                                metadata={"row_number": i + 1}
                            )
                        
                        rows.append(batch_row)
                        
                    except Exception as e:
                        logger.warning(f"Skipping invalid row {i + 1}: {e}")
                        continue
        
        elif upload_request.file_format in [BatchFileFormat.JSON, BatchFileFormat.JSONL]:
            # Parse JSON/JSONL
            with open(file_path, 'r', encoding='utf-8') as f:
                if upload_request.file_format == BatchFileFormat.JSON:
                    data = json.load(f)
                    if not isinstance(data, list):
                        raise ValueError("JSON file must contain an array of objects")
                    json_rows = data
                else:  # JSONL
                    json_rows = []
                    for line_num, line in enumerate(f, 1):
                        line = line.strip()
                        if line:
                            try:
                                json_rows.append(json.loads(line))
                            except json.JSONDecodeError as e:
                                logger.warning(f"Skipping invalid JSON on line {line_num}: {e}")
                
                for i, row_data in enumerate(json_rows):
                    try:
                        if not isinstance(row_data, dict):
                            continue
                        
                        agent_output = row_data.get("agent_output", "").strip()
                        if not agent_output:
                            continue
                        
                        batch_row = BatchInputRow(
                            agent_output=agent_output,
                            ground_truth=row_data.get("ground_truth"),
                            query=row_data.get("query"),
                            agent_id=row_data.get("agent_id"),
                            metadata={**row_data, "row_number": i + 1}
                        )
                        rows.append(batch_row)
                        
                    except Exception as e:
                        logger.warning(f"Skipping invalid row {i + 1}: {e}")
                        continue
        
        return rows
    
    async def start_processing(self, job_id: str) -> bool:
        """Start processing a batch job."""
        if job_id not in self.jobs:
            return False
        
        job = self.jobs[job_id]
        if job.status != BatchStatus.PENDING:
            return False
        
        # Update job status
        job.status = BatchStatus.PROCESSING
        job.started_at = datetime.utcnow()
        
        # Start processing task
        task = asyncio.create_task(self._process_job(job))
        self.active_tasks[job_id] = task
        
        logger.info(f"Started processing batch job {job_id}")
        return True
    
    async def _process_job(self, job: BatchJob):
        """Process a batch job."""
        try:
            # Parse input file
            upload_request = BatchUploadRequest(
                filename=job.filename,
                file_format=job.file_format,
                has_headers=True  # Default assumption
            )
            input_rows = await self._parse_file(Path(job.input_file_path), upload_request)
            
            results = []
            start_time = datetime.utcnow()
            
            for i, input_row in enumerate(input_rows):
                if job.status == BatchStatus.CANCELLED:
                    break
                
                try:
                    # Process single row
                    result = await self._process_single_row(i + 1, input_row)
                    results.append(result)
                    
                    # Update progress
                    job.processed_rows = i + 1
                    if result.error is None:
                        job.successful_rows += 1
                    else:
                        job.failed_rows += 1
                    
                    # Send progress update
                    progress = BatchProgressUpdate(
                        job_id=job.job_id,
                        status=job.status,
                        processed_rows=job.processed_rows,
                        total_rows=job.total_rows,
                        progress_percentage=(job.processed_rows / job.total_rows) * 100,
                        current_row_data=result,
                        estimated_time_remaining=self._estimate_time_remaining(
                            start_time, job.processed_rows, job.total_rows
                        )
                    )
                    await self._notify_progress(progress)
                    
                    # Small delay to prevent overwhelming the API
                    await asyncio.sleep(0.1)
                    
                except Exception as e:
                    logger.error(f"Error processing row {i + 1}: {e}")
                    error_result = BatchResultRow(
                        row_id=i + 1,
                        agent_output=input_row.agent_output,
                        ground_truth=input_row.ground_truth,
                        query=input_row.query,
                        agent_id=input_row.agent_id,
                        hallucination_risk=0.0,
                        flagged=False,
                        confidence=0.0,
                        uncertainty=1.0,
                        processing_time_ms=0.0,
                        error=str(e),
                        metadata=input_row.metadata
                    )
                    results.append(error_result)
                    job.failed_rows += 1
                    job.processed_rows = i + 1
            
            # Save results
            output_file_path = self.storage_dir / f"{job.job_id}_results.json"
            with open(output_file_path, 'w') as f:
                json.dump([result.dict() for result in results], f, indent=2, default=str)
            
            job.output_file_path = str(output_file_path)
            
            # Calculate summary statistics
            successful_results = [r for r in results if r.error is None]
            if successful_results:
                job.average_risk_score = sum(r.hallucination_risk for r in successful_results) / len(successful_results)
                job.flagged_percentage = (sum(1 for r in successful_results if r.flagged) / len(successful_results)) * 100
                job.average_processing_time = sum(r.processing_time_ms for r in successful_results) / len(successful_results)
            
            # Mark job as completed
            job.status = BatchStatus.COMPLETED
            job.completed_at = datetime.utcnow()
            
            # Send final progress update
            final_progress = BatchProgressUpdate(
                job_id=job.job_id,
                status=job.status,
                processed_rows=job.processed_rows,
                total_rows=job.total_rows,
                progress_percentage=100.0
            )
            await self._notify_progress(final_progress)
            
            logger.info(f"Completed batch job {job.job_id}: {job.successful_rows}/{job.total_rows} successful")
            
        except Exception as e:
            logger.error(f"Batch job {job.job_id} failed: {e}")
            job.status = BatchStatus.FAILED
            job.error_message = str(e)
            job.completed_at = datetime.utcnow()
            
            # Send error progress update
            error_progress = BatchProgressUpdate(
                job_id=job.job_id,
                status=job.status,
                processed_rows=job.processed_rows,
                total_rows=job.total_rows,
                progress_percentage=(job.processed_rows / job.total_rows) * 100 if job.total_rows > 0 else 0,
                error_message=str(e)
            )
            await self._notify_progress(error_progress)
        
        finally:
            # Clean up task
            if job.job_id in self.active_tasks:
                del self.active_tasks[job.job_id]
    
    async def _process_single_row(self, row_id: int, input_row: BatchInputRow) -> BatchResultRow:
        """Process a single input row."""
        start_time = datetime.utcnow()
        
        try:
            # Create test request
            test_request = AgentTestRequest(
                agent_output=input_row.agent_output,
                ground_truth=input_row.ground_truth or "Evaluate for factual accuracy and hallucinations based on general knowledge.",
                conversation_history=[]
            )
            
            # Run detection
            detection_result = await self.judge.evaluate(test_request)
            
            # Calculate processing time
            processing_time = (datetime.utcnow() - start_time).total_seconds() * 1000
            
            # Extract detailed results
            claude_score = detection_result.details.get("claude_score")
            statistical_score = detection_result.details.get("statistical_score")
            claude_explanation = detection_result.details.get("claude_explanation", "")
            flagged_segments = detection_result.details.get("hallucinated_segments", [])
            
            return BatchResultRow(
                row_id=row_id,
                agent_output=input_row.agent_output,
                ground_truth=input_row.ground_truth,
                query=input_row.query,
                agent_id=input_row.agent_id,
                hallucination_risk=detection_result.hallucination_risk,
                flagged=detection_result.hallucination_risk > 0.5,
                confidence=1 - detection_result.uncertainty,
                uncertainty=detection_result.uncertainty,
                claude_score=claude_score,
                statistical_score=statistical_score,
                claude_explanation=claude_explanation,
                flagged_segments=flagged_segments,
                processing_time_ms=processing_time,
                metadata=input_row.metadata
            )
            
        except Exception as e:
            processing_time = (datetime.utcnow() - start_time).total_seconds() * 1000
            return BatchResultRow(
                row_id=row_id,
                agent_output=input_row.agent_output,
                ground_truth=input_row.ground_truth,
                query=input_row.query,
                agent_id=input_row.agent_id,
                hallucination_risk=0.0,
                flagged=False,
                confidence=0.0,
                uncertainty=1.0,
                processing_time_ms=processing_time,
                error=str(e),
                metadata=input_row.metadata
            )
    
    def _estimate_time_remaining(self, start_time: datetime, processed: int, total: int) -> Optional[int]:
        """Estimate remaining processing time in seconds."""
        if processed == 0:
            return None
        
        elapsed = (datetime.utcnow() - start_time).total_seconds()
        rate = processed / elapsed  # rows per second
        remaining_rows = total - processed
        
        if rate > 0:
            return int(remaining_rows / rate)
        return None
    
    def get_job(self, job_id: str) -> Optional[BatchJob]:
        """Get job by ID."""
        return self.jobs.get(job_id)
    
    def list_jobs(self, page: int = 1, page_size: int = 20) -> tuple[List[BatchJob], int]:
        """List all jobs with pagination."""
        all_jobs = list(self.jobs.values())
        all_jobs.sort(key=lambda x: x.created_at, reverse=True)
        
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        
        return all_jobs[start_idx:end_idx], len(all_jobs)
    
    async def cancel_job(self, job_id: str) -> bool:
        """Cancel a running job."""
        if job_id not in self.jobs:
            return False
        
        job = self.jobs[job_id]
        if job.status not in [BatchStatus.PENDING, BatchStatus.PROCESSING]:
            return False
        
        job.status = BatchStatus.CANCELLED
        job.completed_at = datetime.utcnow()
        
        # Cancel task if running
        if job_id in self.active_tasks:
            self.active_tasks[job_id].cancel()
        
        logger.info(f"Cancelled batch job {job_id}")
        return True
    
    async def export_results(self, job_id: str, format: str = "csv", 
                           include_metadata: bool = True, 
                           include_explanations: bool = True,
                           filter_flagged_only: bool = False) -> Optional[Path]:
        """Export job results in specified format."""
        job = self.get_job(job_id)
        if not job or job.status != BatchStatus.COMPLETED or not job.output_file_path:
            return None
        
        # Load results
        with open(job.output_file_path, 'r') as f:
            results_data = json.load(f)
        
        results = [BatchResultRow(**data) for data in results_data]
        
        # Filter if requested
        if filter_flagged_only:
            results = [r for r in results if r.flagged]
        
        # Export based on format
        export_path = self.storage_dir / f"{job_id}_export.{format}"
        
        if format == "csv":
            await self._export_csv(results, export_path, include_metadata, include_explanations)
        elif format == "json":
            await self._export_json(results, export_path, include_metadata, include_explanations)
        elif format == "xlsx":
            await self._export_xlsx(results, export_path, include_metadata, include_explanations)
        else:
            return None
        
        return export_path
    
    async def _export_csv(self, results: List[BatchResultRow], path: Path, 
                         include_metadata: bool, include_explanations: bool):
        """Export results to CSV."""
        fieldnames = [
            "row_id", "agent_output", "ground_truth", "query", "agent_id",
            "hallucination_risk", "flagged", "confidence", "uncertainty",
            "processing_time_ms", "error"
        ]
        
        if include_explanations:
            fieldnames.extend(["claude_score", "statistical_score", "claude_explanation", "flagged_segments"])
        
        if include_metadata:
            fieldnames.append("metadata")
        
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for result in results:
                row_data = result.dict()
                if not include_explanations:
                    for key in ["claude_score", "statistical_score", "claude_explanation", "flagged_segments"]:
                        row_data.pop(key, None)
                
                if not include_metadata:
                    row_data.pop("metadata", None)
                else:
                    # Convert metadata to JSON string
                    row_data["metadata"] = json.dumps(row_data.get("metadata", {}))
                
                # Convert lists to strings
                if "flagged_segments" in row_data:
                    row_data["flagged_segments"] = "; ".join(row_data["flagged_segments"])
                
                writer.writerow(row_data)
    
    async def _export_json(self, results: List[BatchResultRow], path: Path, 
                          include_metadata: bool, include_explanations: bool):
        """Export results to JSON."""
        export_data = []
        for result in results:
            row_data = result.dict()
            
            if not include_explanations:
                for key in ["claude_score", "statistical_score", "claude_explanation", "flagged_segments"]:
                    row_data.pop(key, None)
            
            if not include_metadata:
                row_data.pop("metadata", None)
            
            export_data.append(row_data)
        
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, default=str)
    
    async def _export_xlsx(self, results: List[BatchResultRow], path: Path, 
                          include_metadata: bool, include_explanations: bool):
        """Export results to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # Fallback to CSV if openpyxl not available
            csv_path = path.with_suffix('.csv')
            await self._export_csv(results, csv_path, include_metadata, include_explanations)
            return
        
        # Convert to DataFrame for easier Excel export
        data = []
        for result in results:
            row_data = result.dict()
            
            if not include_explanations:
                for key in ["claude_score", "statistical_score", "claude_explanation", "flagged_segments"]:
                    row_data.pop(key, None)
            
            if not include_metadata:
                row_data.pop("metadata", None)
            else:
                row_data["metadata"] = json.dumps(row_data.get("metadata", {}))
            
            # Convert lists to strings
            if "flagged_segments" in row_data:
                row_data["flagged_segments"] = "; ".join(row_data["flagged_segments"])
            
            data.append(row_data)
        
        df = pd.DataFrame(data)
        
        # Save to Excel with formatting
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Results', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Results']
            
            # Style headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # Highlight flagged rows
            flagged_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            for row in worksheet.iter_rows(min_row=2):
                flagged_cell = None
                for cell in row:
                    if worksheet.cell(1, cell.column).value == "flagged":
                        flagged_cell = cell
                        break
                
                if flagged_cell and flagged_cell.value:
                    for cell in row:
                        cell.fill = flagged_fill


# Global batch processor instance
_batch_processor: Optional[BatchProcessor] = None

def get_batch_processor(claude_api_key: str) -> BatchProcessor:
    """Get or create batch processor instance."""
    global _batch_processor
    if _batch_processor is None:
        _batch_processor = BatchProcessor(claude_api_key)
    return _batch_processor
